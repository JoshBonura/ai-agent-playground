# ===== aimodel/file_read/rag/ingest/excel_ingest.py =====
from __future__ import annotations
from typing import Tuple, List
import io, re, hashlib
from datetime import datetime, date, time
from ...core.settings import SETTINGS
from .excel_ingest_core import (
    scan_blocks_by_blank_rows,
    rightmost_nonempty_header,
    select_indices,
)

_WS_RE = re.compile(r"[ \t]+")
def _squeeze_spaces_inline(s: str) -> str:
    return _WS_RE.sub(" ", (s or "")).strip()

def extract_excel(data: bytes) -> Tuple[str, str]:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.utils.datetime import from_excel as _from_excel

    S = SETTINGS.effective

    sig = int(S().get("excel_number_sigfigs"))
    maxp = int(S().get("excel_decimal_max_places"))
    trim = bool(S().get("excel_trim_trailing_zeros"))
    drop_midnight = bool(S().get("excel_dates_drop_time_if_midnight"))
    time_prec = str(S().get("excel_time_precision"))
    max_chars = int(S().get("excel_value_max_chars"))
    quote_strings = bool(S().get("excel_quote_strings"))

    INFER_MAX_ROWS = int(S().get("excel_infer_max_rows"))
    INFER_MAX_COLS = int(S().get("excel_infer_max_cols"))
    INFER_MIN_HEADER_FILL = float(S().get("excel_infer_min_header_fill_ratio", 0.5))
    HEADER_NORMALIZE = bool(S().get("excel_header_normalize"))

    def clip(s: str) -> str:
        if max_chars > 0 and len(s) > max_chars:
            return s[:max_chars] + "…"
        return s

    def fmt_number(v) -> str:
        try:
            s = format(float(v), f".{sig}g") if sig > 0 else f"{float(v):.{maxp}f}"
        except Exception:
            s = str(v)
        if "e" in s.lower():
            try:
                s = f"{float(v):.{maxp}f}"
            except Exception:
                pass
        if trim and "." in s:
            s = s.rstrip("0").rstrip(".")
        return s

    def fmt_date(dt: datetime) -> str:
        if drop_midnight and dt.time() == time(0, 0, 0):
            return dt.date().isoformat()
        return dt.strftime("%Y-%m-%d %H:%M" if time_prec == "minute" else "%Y-%m-%d %H:%M:%S")

    def fmt_time(t: time) -> str:
        return t.strftime("%H:%M" if time_prec == "minute" else "%H:%M:%S")

    def fmt_val(v) -> str:
        if v is None:
            return ""
        if isinstance(v, (int, float)):
            return fmt_number(v)
        if isinstance(v, datetime):
            return fmt_date(v)
        if isinstance(v, date):
            return v.isoformat()
        if isinstance(v, time):
            return fmt_time(v)
        s = str(v)
        if "\n" in s or "\r" in s:
            s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "\\n")
        s = clip(_squeeze_spaces_inline(s))
        if quote_strings and re.search(r"[^A-Za-z0-9_.-]", s):
            return f"\"{s}\""
        return s

    def normalize_header(h: str) -> str:
        if not HEADER_NORMALIZE:
            return h
        s = (h or "").strip().lower()
        s = re.sub(r"[^a-z0-9]+", "_", s)
        s = re.sub(r"_+", "_", s).strip("_")
        return s or h

    def _sheet_used_range(ws: Worksheet):
        from openpyxl.utils import range_boundaries
        if callable(getattr(ws, "calculate_dimension", None)):
            dim_ref = ws.calculate_dimension()
            try:
                min_c, min_r, max_c, max_r = range_boundaries(dim_ref)
                return min_c, min_r, max_c, max_r
            except Exception:
                pass
        return 1, 1, ws.max_column or 1, ws.max_row or 1

    def _cap_rows(min_r: int, max_r: int) -> int:
        return max_r if INFER_MAX_ROWS <= 0 else min(max_r, min_r + INFER_MAX_ROWS - 1)

    def _keep_and_rename_phantom(headers: List[str]) -> tuple[List[int], List[str]]:
        if len(headers) >= 2 and re.fullmatch(r"\d+_\d+$", (headers[1] or "")) and (headers[0] or ""):
            keep_idx = [0, 1]
            new_headers = headers[:]
            new_headers[1] = "value"
            return keep_idx, new_headers
        keep_idx = [i for i, h in enumerate(headers) if h and not re.fullmatch(r"\d+_\d+$", h)]
        new_headers = [headers[i] for i in keep_idx]
        return keep_idx, new_headers

    wb_vals = load_workbook(io.BytesIO(data), data_only=True, read_only=True)

    def _coerce_excel_datetime(cell, v):
        try:
            if getattr(cell, "is_date", False) and isinstance(v, (int, float)):
                return _from_excel(v, wb_vals.epoch)
        except Exception:
            pass
        return v

    lines: List[str] = []
    ingest_id = hashlib.sha1(data).hexdigest()[:16]
    lines.append(f"# Ingest-ID: {ingest_id}")

    def _emit_inferred_table(ws: Worksheet, sheet_name: str, min_c, min_r, max_c, max_r):
        lines.append(f"# Sheet: {sheet_name}")
        max_c_eff = min(max_c, min_c + INFER_MAX_COLS - 1)
        max_r_eff = _cap_rows(min_r, max_r)
        headers: List[str] = []
        header_fill = 0
        for c in range(min_c, max_c_eff + 1):
            val = ws.cell(row=min_r, column=c).value
            s = fmt_val("" if val is None else str(val).strip())
            if s:
                header_fill += 1
            headers.append(s)
        fill_ratio = header_fill / max(1, (max_c_eff - min_c + 1))
        if fill_ratio < INFER_MIN_HEADER_FILL and (min_r + 1) <= max_r_eff:
            headers = []
            hdr_r = min_r + 1
            for c in range(min_c, max_c_eff + 1):
                val = ws.cell(row=hdr_r, column=c).value
                s = fmt_val("" if val is None else str(val).strip())
                headers.append(s)
            min_r = hdr_r
        norm_headers = [normalize_header(h) for h in headers]
        rmax = rightmost_nonempty_header(norm_headers)
        if rmax >= 0:
            norm_headers = norm_headers[:rmax + 1]
            max_c_eff = min(max_c_eff, min_c + rmax)
        keep_idx, norm_headers = _keep_and_rename_phantom(norm_headers)
        lines.append("## Table: " + f"{sheet_name}!R{min_r}-{max_r_eff},C{min_c}-{max_c_eff}")
        if any(h for h in norm_headers):
            lines.append("headers: " + ", ".join(h for h in norm_headers if h))
        lines.append("")
        for r in range(min_r + 1, max_r_eff + 1):
            raw_vals: List[str] = []
            for c in range(min_c, max_c_eff + 1):
                cell = ws.cell(row=r, column=c)
                vv = _coerce_excel_datetime(cell, cell.value)
                raw_vals.append(fmt_val(vv))
            row_vals = select_indices(raw_vals, keep_idx)
            while row_vals and (row_vals[-1] == "" or row_vals[-1] is None):
                row_vals.pop()
            if not any(v for v in row_vals):
                continue
            pairs: List[str] = []
            for h, v in zip(norm_headers, row_vals):
                if h and v:
                    pairs.append(f"{h}={v}")
            lines.append(f"### Row {r} — {sheet_name}")
            lines.append(", ".join(pairs) if pairs else ", ".join(row_vals))
            lines.append("")

    for sheet_name in wb_vals.sheetnames:
        ws_v: Worksheet = wb_vals[sheet_name]
        min_c, min_r, max_c, max_r = _sheet_used_range(ws_v)
        max_c = min(max_c, min_c + INFER_MAX_COLS - 1)
        for b_min_c, b_min_r, b_max_c, b_max_r in scan_blocks_by_blank_rows(ws_v, min_c, min_r, max_c, max_r):
            if b_min_r > b_max_r:
                continue
            _emit_inferred_table(ws_v, sheet_name, b_min_c, b_min_r, b_max_c, b_max_r)

    text = "\n".join(line.rstrip() for line in lines if line is not None).strip()
    return (text + "\n" if text else ""), "text/plain"
